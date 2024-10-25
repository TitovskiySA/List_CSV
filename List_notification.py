#!/usr/bin/env python
#-*- coding: utf-8 -*-

import sys
import os
import wx
#import wx.grid
#from wx.adv import SplashScreen as SplashScreen
import time
import datetime  # импорт библиотеки дат и времени
from datetime import datetime, timedelta
import locale
import csv # CSV library
import openpyxl # Excel libraries
from openpyxl import Workbook 
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from openpyxl.utils.cell import get_column_letter

#========================================================================================
# создание класса основного окна
class OsnWindow(wx.Frame):

    # задаем конструктор
    def __init__(self, parent):

        # создаем стиль окна без кнопок закрытия и тд
        styleWindow = (
            wx.MINIMIZE_BOX|
            wx.MAXIMIZE_BOX|
            wx.RESIZE_BORDER|
            wx.CAPTION|
            wx.SYSTEM_MENU|
            wx.CLOSE_BOX|
            wx.CLIP_CHILDREN
            )

        #задаем обращение к параметрам родительского класса (чтобы не создавать их заново)
        super().__init__(
            parent,
            title = "CSV converter ver. 1.0",
            #size = (500, 200),
            style = styleWindow)
        #------------------------------------------------------------------------------
        #задаем иконку
        frameIcon = wx.Icon(os.getcwd() + "\\images\\comfolder.ico")
        self.SetIcon(frameIcon)

        panel = MainPanel(self)
        self.Center()
        #self.Show()
        
# panel
class MainPanel(wx.Panel):

    def __init__(self, parent):
        wx.Panel.__init__(self, parent = parent)
        self.frame = parent
        self.MyDir = os.getcwd()
        # Создание vbox 
        CommonVbox = wx.BoxSizer(wx.VERTICAL)
        #-----------------------------------------------------'''
        # Добавление меню
        self.menuBar = wx.MenuBar()      
        MenuAbout = wx.Menu()
        self.menuBar.Append(MenuAbout, "Справка")
        self.frame.SetMenuBar(self.menuBar)
        #-----------------------------------------------------'''
        # Заполнение меню Справка        
        License = MenuAbout.Append(-1, "Лицензия")
        self.frame.Bind(wx.EVT_MENU, self.ShowLic, License)

        AboutPO = MenuAbout.Append(-1, "Версия")
        self.frame.Bind(wx.EVT_MENU, self.ShowInfo, AboutPO)
        #-----------------------------------------------------'''
        # Buttons
        CSVRaw = wx.Button(self, wx.ID_ANY, "Справка для системы оповещения")
        CSVRaw.Bind(wx.EVT_BUTTON, self.SpisokXLSX)
        CSVRaw.SetFont(wx.Font(14, wx.ROMAN, wx.NORMAL, wx.NORMAL))
        CommonVbox.Add(CSVRaw, 1, wx.ALL|wx.EXPAND, 20)

        #-----------------------------------------------------------
        #Hotkeys
        RawCommand = wx.NewIdRef()
        #ATSCommand = wx.NewIdRef()
        self.frame.Bind(wx.EVT_MENU, self.SpisokXLSX, id = RawCommand)
        #self.frame.Bind(wx.EVT_MENU, self.CSVtoXLSX, id = ATSCommand)
        entries = [wx.AcceleratorEntry() for i in range(1)]
        entries[0].Set(wx.ACCEL_CTRL|wx.ACCEL_SHIFT, ord("R"), RawCommand)
        #entries[1].Set(wx.ACCEL_CTRL|wx.ACCEL_SHIFT, ord("A"), ATSCommand)

        accel = wx.AcceleratorTable(entries)
        self.frame.SetAcceleratorTable(accel)

        self.Log = ""
        # Создание папок и файлов
        try:
            os.chdir(self.MyDir)
            os.mkdir("logs")
        except Exception as Err:
            ToLog("Can't create log folder because of = " + str(Err))
        
        # Действия при нажатии кнопки закрыть
        self.frame.Bind(wx.EVT_CLOSE, self.OnCloseWindow)

        self.SetSizer(CommonVbox)
        self.Fit()

        # делаем окно видимым
        self.frame.Show(True)
    
#=======================================================================================
    # При нажатии кнопки закрыть
    def OnCloseWindow(self, event):
        print ("Нажато закрыть")

        dlg = wx.MessageDialog(
            self,
            "Вы уверены, что хотите выйти из программы?",
            "Выход",
            wx.OK|wx.CANCEL)

        # Действия при нажатии кнопки OK
        if dlg.ShowModal() != wx.ID_OK:
            return
        else:
            ToLog("Application closed")
            ToLog("\n----------------------------\n\n")
            #self.Destroy()
            wx.Exit()
            sys.exit()
                   
        return
#===========================================================================================
    # Лицензия
    def ShowLic(self, event):

        LICENSE = (
            "Данная программа является свободным программным обеспечением\n"+
            "Вы вправе распространять её и/или модифицировать в соответствии\n"+
            "с условиями версии 2 либо по Вашему выбору с условиями более\n"+
            "поздней версии Стандартной общественной лицензии GNU, \n"+
            "опубликованной Free Software Foundation.\n\n\n"+
            "Эта программа создана в надежде, что будет Вам полезной, однако\n"+
            "на неё нет НИКАКИХ гарантий, в том числе гарантии товарного\n"+
            "состояния при продаже и пригодности для использования в\n"+
            "конкретных целях.\n"+
            "Для получения более подробной информации ознакомьтесь со \n"+
            "Стандартной Общественной Лицензией GNU.\n\n"+
            "Данная программа написана на Python\n\n"
            "Автор: Титовский С.А.\n")
        
        # Создание диалогового окна
        dlglic = wx.MessageBox(LICENSE, "Лицензия", wx.OK)
        return
#==============================================================================================
    def ShowInfo(self, event):
        print ("Show Info")
        try:
            path = os.path.realpath(os.getcwd() + "\\info")
            os.startfile(path + "\\AboutServer.pdf")
            
        except Exception as Err:
            wx.MessageBox("Версия 0.2 от 24.08.2023", " ", wx.OK)
            #self.ToLog("Ошибка показа справки с кодом = " + str(Err))
        return
    
#=============================================================================================
    def SpisokXLSX(self, event):
        try:
            self.DataCSV = self.ConvertToList()
            ToLog("1. Converting CSV to XLSX successed")
        except Exception as Err:
            ToLog("!!! Error Convert CSV, Error code = " + str(Err))
            wx.MessageBox("Ошибка обработки файла .CSV", " ", wx.OK)
            return
 
        self.DoSpisokXLSX(self.DataCSV)   
          
#================================================================================
    def ConvertToList(self):
        DialogLoad = wx.FileDialog(
            self,
            "Выберите файл для преобразования",
            #defaultDir = os.getcwd(),
            wildcard = "CSV files (*.csv)|*csv",
            style = wx.FD_OPEN)
        if DialogLoad.ShowModal() == wx.ID_CANCEL:
            return
        self.LoadDir = DialogLoad.GetDirectory()
        if ".csv" in DialogLoad.GetFilename():
            self.LoadFile = DialogLoad.GetFilename()
        else:
            self.LoadFile = DialogLoad.GetFilename() + ".csv"
        ToLog("Для преобразования выбран .csv файл: " + self.LoadDir + "\\" + self.LoadFile)

        # work with csv
        csvdata = []
        with open(self.LoadDir + "\\" + self.LoadFile, "r") as csvfile:
            csvreader = csv.reader(csvfile)
            for row in csvreader:
                #print(str(row))
                if row != []:
                    for i in range(17,0,-1):
                        try:
                            row[0] = row[0].replace((i + 1)*";",";")
                        except Exception:
                            pass
                    csvdata.append(row[0])
                    #print(row[0])
                else:
                    csvdata.append("PUSTO")

        Data = []
        for row in csvdata:
            Data.append(row)
            #print(str(row))

        return Data
#===================================================================================
    def StylesXLSX(self):
        # Fonts
        self.FontHeadBold = Font(name = "Times New Roman", size = 14, bold = True)
        self.FontHead = Font(name = "Times New Roman", size = 14, bold = False)
        self.FontTableBold = Font(name = "Times New Roman", size = 12, bold = True)
        self.FontTable = Font(name = "Times New Roman", size = 12, bold = False)
        
        # Borders
        self.bordertable = Border(left=Side(border_style="thin", color="FF000000"),
                                  right=Side(border_style="thin", color="FF000000"),
                                  top=Side(border_style="thin", color="FF000000"),
                                  bottom=Side(border_style="thin", color="FF000000"))

        # Alignment
        self.aligntable = Alignment(horizontal = "center",
                                    vertical = "center",
                                    text_rotation = 0,
                                    wrap_text = True,
                                    shrink_to_fit = False,
                                    indent = 0)
        
        self.alignhead = Alignment(horizontal = "left",
                                    vertical = "center",
                                    text_rotation = 0,
                                    wrap_text = True,
                                    shrink_to_fit = False,
                                    indent = 0)

        self.aligntable2 = Alignment(horizontal = "center",
                                    vertical = "center",
                                    text_rotation = 90,
                                    wrap_text = True,
                                    shrink_to_fit = False,
                                    indent = 0)
        return
#====================================================================================
#    def ConvertXLSX(self, csvdata):
#        self.StylesXLSX()
        
        # Creating Excel file    
#        self.wb = openpyxl.Workbook()
#        work = self.wb[self.wb.sheetnames[0]]

#        datatable = []
#        temp = []
#        for i in range (0, len(csvdata)):
            #if csvdata[i] == "PUSTO":
            #    continue
        
#            try:
#                if "Начало сеанса" in csvdata[i] and "Завершение вызова" in csvdata[i]:
#                    if ";" in csvdata[i-1]:
#                        csvdata[i-1] = csvdata[i-1].split(";")
#                        csvdata[i-1].insert(3, " ")
#                        datatable[i-1] = csvdata[i-1][:]
#                    #datatable.append(csvdata[i])
#                    datatable.append(csvdata[i].split(";"))
#                else:
#                    #datatable.append(csvdata[i])
#                    datatable.append(csvdata[i].split(";"))
#            except Exception:
#                #datatable.append(csvdata[i])
#                datatable.append(csvdata[i].split(";"))
#                print("Added with Error" + str(csvdata[i]))

#        for i in range (0 , len(datatable)):
#            if datatable[i] == ["PUSTO"] or (i > 17 and len(datatable[i]) == 2) or (i > 17 and len(datatable[i]) == 3):
#                continue
#            else:
#                temp.append(datatable[i][:])
#        datatable = temp[:]
                    
#        for i in range (0, len(datatable)):
#            #print(str(datatable[i]))
#            if len(datatable[i]) == "PUSTO":
#                work.cell(row = 1 + i, column = 1, value = "").font=self.fonttable
#            elif isinstance (datatable[i], list):
#                for ii in range (0, len(datatable[i])):
#                    work.cell(row = 1 + i, column = 1 + ii, value = str(datatable[i][ii])).font=self.fonttable                  
#            elif ";" not in datatable[i]:
#                work.cell(row = 1 + i, column = 1, value = str(datatable[i])).font=self.fonttable
#            else:
#                datatable[i] = datatable[i].split(";")
#                for ii in range (0, len(datatable[i])):
#                    work.cell(row = 1 + i, column = 1 + ii, value = str(datatable[i][ii])).font=self.fonttable

#        dimensions = [12,25,25,25,16,25,25,16,24]
#        for i in range (0, len(dimensions)):
#            work.column_dimensions[get_column_letter(i + 1)].width = dimensions[i]
#        return

#========================================================================================================
    def DoSpisokXLSX(self, data):
        try:
            self.DataToListOne = DoListOne(data)
            ToLog("2. Converting Data from XLSX successed")
            self.FillTable1()
            ToLog("3. Filling Table1 XLSX with data successed")
            self.FillTable2()
            ToLog("4. Filling Table12 XLSX with data successed")
            self.SavingExcel(self.wb)
            ToLog("5. Saving XLSX file successed")
        except Exception as Err:
            ToLog("!!! Error DoSpisokXLSX function, Error code = " + str(Err))
            SomeError(
                "Ошибка",
                "При обработке файла .CSV произошла ошибка" +
                "\nКод ошибки:\n" + str(Err))
            #raise Exception               
        return

#=====================================================================================================
    def FillTable1(self):
        # Creating Excel file    
        self.wb = openpyxl.Workbook()
        self.wb.create_sheet("Список оповещения")
        self.wb.remove(self.wb[self.wb.sheetnames[0]])
        work = self.wb[self.wb.sheetnames[0]]
        work.page_setup.orientation = work.ORIENTATION_LANDSCAPE
        work.page_setup.fitToPage = True
        work.page_setup.fitToHeight = False

        #fonts
        self.StylesXLSX()

        #names 
        HeadDoc = [
            "Отчёт по оповещению должностных лиц по Управлению Западно-Сибирской железной дороги",
            "Сигнал",
            self.DataToListOne[6][3],
            "поступил от",
            "Сигнал передан старшему смены НС:",
            str(datetime.today())[8:10] + "." + str(datetime.today())[5:7] + "." +
            str(datetime.today())[0:4] + " в 00:00 (мск.вр.)"]

        HeadTable = [
            "№",
            "Оповещаемые филиалы," + "\n" + "структурные подразделения и ДЗО",
            "Телефон",
            "Ф.И.О.",
            "Время" + "\n" + "передачи" + "\n" + "сигнала"]

        EndTable = [
            "Итого",
            "Количество оповещаемых",
            "Время начала\nоповещения",
            "Время окончания\nоповещения",
            "Общее время на оповещение"]

        EndDoc = [
            "Начало оповещения",
            "Конец оповещения",
            "Старший смены ЦТУ________________________",
            "Начальник специальной службы________________________"]

        
        #merge cells
        work.merge_cells("A2:F2")
        work.merge_cells("B4:E4")
        work.merge_cells("A5:B5")
        work.merge_cells("D5:E5")
        work.merge_cells("A6:B6")
        work.merge_cells("D6:E6")
        colour = "FFFF00"

        #Filling Head Doc
        work.cell(row = 2, column = 1, value = HeadDoc[0]).font = self.FontHead
        work.cell(row = 4, column = 1, value = HeadDoc[1]).font = self.FontHead
        work.cell(row = 4, column = 2, value = HeadDoc[2]).font = self.FontHeadBold
        work.cell(row = 5, column = 1, value = HeadDoc[3]).font = self.FontHead
        work.cell(row = 5, column = 1).alignment = self.alignhead
        work.cell(row = 6, column = 1, value = HeadDoc[4]).font = self.FontHead
        work.cell(row = 6, column = 1).alignment = self.alignhead
        work.cell(row = 6, column = 4, value = HeadDoc[5]).font = self.FontHead
        work.cell(row = 6, column = 4).alignment = self.alignhead
        work.cell(row = 6, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        

        #Filling HeadTable
        for Col in range (0, len(HeadTable)):
            work.cell(row = 8, column = Col + 1, value = HeadTable[Col]).font = self.FontTableBold

        #Filling Data in Table
        for Col in range (0, len(self.DataToListOne) - 2):
            for Row in range (0, len(self.DataToListOne[Col])):
                work.cell(row = Row + 9, column = Col + 1, value = self.DataToListOne[Col][Row]).font = self.FontTable
                if (
                    (Col == 4 and self.DataToListOne[Col][Row] == "Уточнить в РМТС")
                    or
                    (Col == 4 and self.DataToListOne[Col][Row] == "Выключен")
                    or
                    (Col == 1 and self.DataToListOne[Col][Row].isdigit() == True)):
                    work.cell(row = Row + 9, column = Col + 1).fill = PatternFill(
                        fgColor = colour, fill_type = "solid")

        RowEnd = len(self.DataToListOne[0]) + 1 + 7
        
        #Filling EndTable
        work.merge_cells("A" + str(RowEnd + 1) + ":A" + str(RowEnd + 2))
        for Col in range (0, len(EndTable)):
            work.cell(row = RowEnd + 1, column = Col + 1, value = EndTable[Col]).font = self.FontTable
        work.cell(row = RowEnd + 2, column = 2, value = len(self.DataToListOne[0])).font = self.FontTable
        work.cell(row = RowEnd + 2, column = 3, value = self.DataToListOne[6][0]).font = self.FontTable
        work.cell(row = RowEnd + 2, column = 4, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = RowEnd + 2, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = RowEnd + 2, column = 5, value = self.DataToListOne[6][2]).font = self.FontTable
        work.cell(row = RowEnd + 2, column = 5).fill = PatternFill(
            fgColor = colour, fill_type = "solid")

        #Filling End Doc
        work.cell(row = RowEnd + 4, column = 1, value = EndDoc[0]).font = self.FontHead
        work.cell(row = RowEnd + 4, column = 3, value = self.DataToListOne[6][0]).font = self.FontHead
        work.cell(row = RowEnd + 5, column = 1, value = EndDoc[1]).font = self.FontHead
        work.cell(row = RowEnd + 5, column = 3, value = self.DataToListOne[6][1]).font = self.FontHead
        work.cell(row = RowEnd + 5, column = 3).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = RowEnd + 7, column = 3, value = EndDoc[2]).font = self.FontHead
        work.cell(row = RowEnd + 8, column = 3, value = EndDoc[3]).font = self.FontHead

        # make dimensions and alignment
        dimensions = [10, 41, 20, 45, 30]
        rowdimensions = 20

        for Row in range (1, 7):
            work.row_dimensions[Row].height = rowdimensions
            
        for i in range (0, len(dimensions)):
            work.column_dimensions[get_column_letter(i + 1)].width = dimensions[i]
            for ii in range(0, len(self.DataToListOne[0]) + 10):
                if ii >= 7:
                    work.cell(row = ii + 1, column = i + 1).border = self.bordertable
                work.cell(row = ii + 1, column = i + 1).alignment = self.aligntable
               
        return

#=====================================================================================================
    def FillTable2(self):
        # Creating Excel file    
        self.wb.create_sheet("Сводная")
        work = self.wb[self.wb.sheetnames[1]]
        work.page_setup.orientation = work.ORIENTATION_PORTRAIT
        work.page_setup.fitToPage = True
        work.page_setup.fitToHeight = False

        #fonts
        self.StylesXLSX()

        #names 
        HeadDoc = "Отчёт по оповещению должностных лиц" + "\nЗападно-Сибирской железной дороги"
        HeadTable = [
            "Сигнал/приказ\n\n" + "дата: " +
            str(datetime.today())[8:10] + "." + str(datetime.today())[5:7] + "." +
            str(datetime.today())[0:4] + 
            self.DataToListOne[6][3] + "\nвремя: 00:00 (мск.вр.)" +
            "\nполучен от: ДОЛЖНОСТЬ, ФИО",
            "Управление дороги",
            "Омский регион", "Новосибирский регион",
            "Кузбасский регион", "Алтайский регион",
            "Всего по дороге"]
        ColTable = [
            "Время получения",
            "Время начала оповещения в автоматическом режиме",
            "Время завершения оповещения",
            "Время начала оповещения в ручном режиме",
            "Время завершения оповещения",
            "Подлежит оповещению",
            "Оповещено в автоматическом режиме",
            "Оповещено в ручном режиме",
            "Не оповещено"]
        EndDoc = "Старший смены ЦТО _______________"

        #merge cells
        work.merge_cells("A1:I1")
        work.merge_cells("A3:C3")
        work.merge_cells("A4:C4")
        work.merge_cells("A5:C5")
        work.merge_cells("A6:C6")
        work.merge_cells("A7:C7")
        work.merge_cells("A8:C8")
        work.merge_cells("A9:C9")
        work.merge_cells("A10:C10")
        work.merge_cells("A11:C11")
        work.merge_cells("A12:C12")
        work.merge_cells("A14:I14")
        colour = "FFFF00"

        #Filling Head Doc
        work.cell(row = 1, column = 1, value = HeadDoc).font = self.FontHead
        work.cell(row = 1, column = 1).alignment = self.aligntable

        #Filling Head Table
        work.cell(row = 3, column = 1, value = str(HeadTable[0])).font = self.FontTable
        work.cell(row = 3, column = 1).alignment = self.aligntable
        work.cell(row = 3, column = 1).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        for col in range (1, len(HeadTable)):
            work.cell(row = 3, column = 3 + col, value = str(HeadTable[col])).font = self.FontTable
            work.cell(row = 3, column = 3 + col).alignment = self.aligntable2
            work.cell(row = 3, column = 3 + col).border = self.bordertable
        work.cell(row = 3, column = 1).border = self.bordertable
        work.cell(row = 3, column = 2).border = self.bordertable
        work.cell(row = 3, column = 3).border = self.bordertable

        #Filling Data in Table
        #FirstCol
        for Row in range (0, len(ColTable)):
            work.cell(row = 4 + Row, column = 1, value = str(ColTable[Row])).font = self.FontTable
            work.cell(row = 4 + Row, column = 1).alignment = self.alignhead
            for Col in range (0, 3):
                work.cell(row = 4 + Row, column = 1 + Col).border = self.bordertable

        #SecondCol
        for Row in range (0, len(ColTable)):
            for Col in range (4, 10):
                work.cell(row = 4 + Row, column = Col).alignment = self.aligntable
                work.cell(row = 4 + Row, column = Col).border = self.bordertable
                
        work.cell(row = 4, column = 4, value = "00:00").font = self.FontTable
        work.cell(row = 4, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        
        work.cell(row = 5, column = 4, value = self.DataToListOne[6][0]).font = self.FontTable
        work.cell(row = 6, column = 4, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = 7, column = 4, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = 7, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 8, column = 4, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = 8, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 9, column = 4, value = len(self.DataToListOne[0])).font = self.FontTable

        #count authomatic
        NonAuthomatic = self.DataToListOne[4].count("Уточнить в РМТС") + self.DataToListOne[4].count("Выключен")
        #print(str(self.DataToListOne[4]))
        work.cell(row = 10, column = 4, value = len(self.DataToListOne[4]) - NonAuthomatic).font = self.FontTable
        work.cell(row = 11, column = 4, value = self.DataToListOne[4].count("Уточнить в РМТС")).font = self.FontTable
        work.cell(row = 11, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 12, column = 4, value = 0).font = self.FontTable
        work.cell(row = 12, column = 4).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        
        #NinthCol
        work.cell(row = 4, column = 9, value = "00:00").font = self.FontTable
        work.cell(row = 4, column = 9).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 5, column = 9, value = self.DataToListOne[6][0]).font = self.FontTable
        work.cell(row = 6, column = 9, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = 7, column = 9, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = 7, column = 9).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 8, column = 9, value = self.DataToListOne[6][1]).font = self.FontTable
        work.cell(row = 8, column = 9).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 9, column = 9, value = "=SUM(D9:H9)").font = self.FontTable
        work.cell(row = 10, column = 9, value = "=SUM(D10:H10)").font = self.FontTable
        work.cell(row = 11, column = 9, value = "=SUM(D11:H11)").font = self.FontTable
        work.cell(row = 11, column = 9).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        work.cell(row = 12, column = 9, value = "=SUM(D12:H12)").font = self.FontTable
        work.cell(row = 12, column = 9).fill = PatternFill(
            fgColor = colour, fill_type = "solid")
        
        #Filling EndTable
        work.cell(row = 14, column = 1, value = EndDoc).font = self.FontTable
        
        # make dimensions and alignment
        coldimensions = [6, 16, 8.5, 8.5, 8.5, 8.5, 8.5, 8.5, 8.5, 8.5]
        rowdimensions = [54, 16, 114, 36, 36, 36, 36, 36, 36, 36, 36, 36, 16, 16]

        for Row in range (0, len(rowdimensions)):
            work.row_dimensions[Row + 1].height = rowdimensions[Row]
            
        for Col in range (0, len(coldimensions)):
            work.column_dimensions[get_column_letter(Col + 1)].width = coldimensions[Col]
                  
        return

#=====================================================================================================
    def SavingExcel(self, wb):
        # Сохранение файла
        DialogSave = wx.FileDialog(
            self,
            "Выберите файл для сохранения",
            #defaultDir = os.getcwd(),
            wildcard = "xlsx files (*.xlsx)|*xlsx",
            style = wx.FD_SAVE | wx.FD_OVERWRITE_PROMPT)
        if DialogSave.ShowModal() == wx.ID_CANCEL:
            return
        self.SaveDir = DialogSave.GetDirectory()
        self.SaveFile = DialogSave.GetFilename()
        ToLog("Для сохранения выбран .xslx файл: " + self.SaveDir + "\\" + self.SaveFile)
        
        try:
            if ".xlsx" in DialogSave.GetFilename():
                self.SaveFile = DialogSave.GetFilename()
            else:
                self.SaveFile = DialogSave.GetFilename() + ".xlsx"
            wb.save(self.SaveDir + "\\" + self.SaveFile)
            ToLog("Файл " + self.SaveDir + "\\" + self.SaveFile + " успешно сохранен")
            # opening xlsx
            try:
                path = os.path.realpath(self.SaveDir + "\\" + self.SaveFile)
                os.startfile(path)
            
            except Exception as Err:
                wx.MessageBox("Не удалось открыть полученный файл," + 
                                "будет открыта папка с сохраненным файлом", " ", wx.OK)
                path = os.path.realpath(self.SaveDir)
                os.startfile(path)
                ToLog("!!! Error SavingExcel, Error code = " + str(Err))
                    
        except Exception as Err:
            ToLog("!!! Error SavingExcel2, Error code = " + str(Err))

        return

#========================================================
#========================================================
#========================================================
#========================================================
def DoListOne(data):
    #DataToList1 = [ [№],[Dolgn], [TelNumber1], [FIO], [time], [timebegin, timeend]]
    #DataToList1 = [["№"], ["Оповещаемые филиалы,\nструктурные подразделения и ДЗО"], ["Телефон"],
    #["Ф.И.О."], ["Время\nпередачи\nсигнала"], ["Результат оповещения"],[timebegin, timeend, timediff, name]]
    DataToList1 = [
        [],[],[],[],[],[],
        ["00:00", "00:00", "00:00", "Сбор оперативного штаба и секретариата по предупреждению и ликвидации чрезвычайных ситуаций"]]
    RawResult = []
    RawTel = []
    RawTime = []
    temp = []
    for row in data:
        row = row.split(";")
        if len(row) == 2 or ["Попыток"] in row or row == ["PUSTO"]:
            continue
        #print(str(row))
        temp.append(row)
    data = temp
    print("here's data")
    #for dat in data:
    #    print(str(dat))
        
    for i in range (1, len(data)):
        #appending timebegin, timeend, name
        if "Запуск" in data[i] and "Завершение" in data[i+1]:
            DataToList1[6][0] = ChangeTime(str(data[i][1].split()[1][:-3]), 4)
            DataToList1[6][1] = ChangeTime(str(data[i+1][1].split()[1][:-3]), 4)
            #DataToList1[6][2] = ChangeTime(str(data[i+2][1][:-3]), 4)
            DataToList1[6][2] = str(str(datetime.strptime(DataToList1[6][1], "%H:%M") - datetime.strptime(DataToList1[6][0], "%H:%M"))[:-3])
            if len(DataToList1[6][2]) == 4:
                DataToList1[6][2] = "0" + DataToList1[6][2]
                
            
            DataToList1[6][3] = str(data[i-2][1])
            print(str(DataToList1[6]))
        
        #appending fio and dolgnost
        if "Начало сеанса" in data[i] and "Завершение вызова" in data[i]:
            DataToList1[1].append(str(data[i-1][2]))
            DataToList1[3].append(str(data[i-1][1]))

            #if vacancy
            if "Не заданы телефоны" in data[i+1]:
                RawResult.append(["Вакансия"])
                RawTel.append(["Вакансия"])
                RawTime.append(["Вакансия"])
                continue

            #if off
            if "Выключен" in data[i+1]:
                RawResult.append(["Выключен"])
                RawTel.append(["Выключен"])
                RawTime.append(["Выключен"])
                continue
                  
            #appending Raw result and tels of calling
            start = i + 1
            result = []
            tel = []
            time = []
            #print("Start = " + str(i))
            while "DTMF" in data[start]:
                result.append(str(data[start][7]))
                tel.append(str(data[start][6]))
                time.append(str(data[start][3].split()[1][:-3]))
                #print("result now = " + str(result))
                if start < len(data) - 1:
                    start = start + 1
                else:
                    break
            RawResult.append(result[:])
            RawTel.append(tel[:])
            RawTime.append(time[:])
  
    #analyzing raw results
    for i in range(0, len(RawResult)):
       # print("number = " + str(i))
        if "Оповещён" in RawResult[i]:
            for ii in range (0, len(RawResult[i])):
                if RawResult[i][ii] == "Оповещён":
                    DataToList1[4].append(ChangeTime(RawTime[i][ii], 4))
                    DataToList1[5].append("Оповещён")
                    DataToList1[2].append(RawTel[i][ii])
                    continue
        elif "Вакансия" in RawResult[i]:
            DataToList1[4].append("Вакансия")
            DataToList1[5].append("Вакансия")
            DataToList1[2].append("Вакансия")
        elif "Выключен" in RawResult[i]:
            DataToList1[4].append("Выключен")
            DataToList1[5].append("Выключен")
            DataToList1[2].append("Выключен")
        else:
            DataToList1[4].append("Уточнить в РМТС")
            DataToList1[5].append("Оповещение в РМТС")
            DataToList1[2].append(RawTel[i][0])
            
            
        #DataToList1[2].append("Change_it")
        DataToList1[0].append(str(i + 1))

    #printing result
    print("#\n#\n#\n#\n#\n#\n#\n#\n#\n#\n#\n#")
    #for dat in DataToList1:
    #    print(str(dat))

    

    #analyzing tel number
    #for i in range (0, len(RawTel)):
        #print("RawTel" + str(i) + " = " + str(RawTel[i]))
        #if call wasnot success
    #    if DataToList1[6][i] == "Уточнить в РМТС":
    #        DataToList1[2][i] = str(RawTel[i][1])
    #        continue
        #if called successed on sot number
        #print("DataToList = " + str(DataToList1[3][i]))
    #    if len(DataToList1[3][i]) == 12:
    #        if len(RawTel[i]) > 1:
    #            for ii in range (0, len(RawTel[i])):
    #                #find working tel
    #                if len(RawTel[i][ii]) < 12:
    #                    DataToList1[2][i] = RawTel[i][ii]
    #                    break
    #            continue
    #        else:
    #            DataToList1[2][i] == "УТОЧНИТЬ 2 НОМЕР"
    #            continue
        #if called successed on work number
    #    else:
    #        if len(RawTel[i]) > 1:
    #            for ii in range (0, len(RawTel[i])):
    #                #find sot tel
    #                if len(RawTel[i][ii]) == 12:
    #                    DataToList1[2][i] = RawTel[i][ii]
    #                    break
    #            continue
    #        else:
    #            DataToList1[2][i] == "УТОЧНИТЬ 2 НОМЕР"
    #            continue
                

    #printing result
    #print("#\n#\n#\n#\n#\n#\n#\n#\n#\n#\n#\n#")
    #for dat in DataToList1:
    #    print(str(dat))

    return DataToList1
#===============================================
#===============================================
#===============================================
#===============================================
#ChangeTimeFunction
def ChangeTime(Time, ChHour = 4):
    try:
        result = str(datetime.strptime(Time, "%H:%M") - timedelta(hours = ChHour)).split()[1][:-3]
    except Exception as Err:
        ToLog("Error in ChangeTimeFunction, Error code = " + str(Err))
        result = Time + "мест.вр."
    return result

#===============================================
#===============================================
#===============================================
#===============================================        
# Создание класса окна любой ошибки
def SomeError(parent, title):
    wx.MessageBox(title, "Ошибка", wx.OK)
#=============================================
#=============================================
#=============================================
#=============================================
# Tolog - renew log
def ToLog(message):
    global MyDir
    try:
        file = open(MyDir + "\\Logs\\" + str(datetime.today())[0:10] + ".txt", "a")
        file.write(str(datetime.today())[10:19] + "  " + str(message) + "\n")
        file.close()
    except Exception as Err:
        SomeError(None, "Ошибка логирования, причина = " + str(Err))
    return
#=============================================
#=============================================
#=============================================
#=============================================
# CreateFolders
def CreateFolders():
    global MyDir
    try:
        os.mkdir(MyDir + "\\Logs")
    except Exception as Err:
        print("Error creating folder of Logs, Error code = " + str(Err))
#=============================================
#=============================================
#=============================================
#=============================================
# ClearOldLogs
def ClearLogs():
    global MyDir
    try:
        print("len Logs = " + str(len(os.listdir(MyDir + "\\Logs"))))
        while len(os.listdir(MyDir + "\\Logs")) >= 10:
            if len(os.listdir(MyDir + "\\Logs")) < 10:
                    break
            try:
                os.remove(os.path.abspath(FindOldest(MyDir + "\\Logs")))
                ToLog("DELETING FILE " + str(FindOldest(MyDir + "\\Logs")))
            except Exception as Err:
                ToLog("Старый файл лога не был удален, код ошибки = " + str(Err))
                #raise Exception
                break
    except Exception as Err:
        ToLog("Ошибка в выполнении функции очитски файла логов, код ошибки = " + str(Err))
        #raise Exception
        
#=============================================
#=============================================
#=============================================
#=============================================   
# DeleteOldest
def FindOldest(Dir):
    try:
        List = os.listdir(Dir)
        fullPath = [Dir + "/{0}".format(x) for x in List]
        oldestFile = min(fullPath, key = os.path.getctime)
        return oldestFile
    except Exception as Err:
        ToLog("Ошибка в поиске старого файла, код ошибки = " + str(Err))
        #raise Exception
        return False
                    
'''============================================================================'''
# Определение локали!
locale.setlocale(locale.LC_ALL, "")

global MyDir
MyDir = os.getcwd()
CreateFolders()
ClearLogs()

ex = wx.App()
ToLog("\n----------------------------")
ToLog("Application Started")
OsnWindow(None)

ex.MainLoop()


















